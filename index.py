#!/home/kels/Documents/projects/netAutomation/readExcelAssest/readExcel/bin/python
#the above sources library from virtual environment
from operator import eq
import sys,os
import time

#Dependancies LIB
import openpyxl as lib
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

#test_acl.xlsx

def genDataTemplate():
    return {"sources":[],"devices":[],"destinations":[],"ports":[]}

def removeNewLine(string):
    return str(string).strip()

def main():
    data=[genDataTemplate()]
    indexControllers=["source","device","destination","port"] #controls which dictionary to add cell data to
    wb = lib.load_workbook("./test_acl.xlsx",read_only=False)
    sheet = wb.worksheets[0]
    x=sheet.max_column
    y=sheet.max_row
    print("No. Rows: %s No. Col %s" %(y,x))
    for rowList in sheet.iter_rows(): #loop through row elements
        for xIndex in range(0,x): #loop every colomn in the row
            if rowList[xIndex].internal_value:
                values=list(map(removeNewLine,str(rowList[xIndex].internal_value).split(',')))
                data[len(data)-1][indexControllers[xIndex]]=values
            # else:
            #     data[len(data)-1][indexControllers[xIndex]]=None
        #done with row, create append a new dict to data
        data.append(genDataTemplate()) #remove first row
    data.pop(0)
    print(data)
    return

if __name__ == '__main__':
    #sys.argv = ["programName.py","--input","test.txt","--output","tmp/test.txt"]
    main()